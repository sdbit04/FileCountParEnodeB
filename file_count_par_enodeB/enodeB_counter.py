import pathlib
import re
import openpyxl
import os


def search_enodeb_n_file_count_per_day(stats_file, stat_date, enode_pattern, enodeb_count_dict):
    """Not in use.
    This function work for one file.
    read, search for particular pattern, increases match count,
    Finally contribute to the storage dict object"""
    stat_date = stat_date
    search_pattern = enode_pattern
    search_pattern_ob = re.compile(search_pattern, re.IGNORECASE)
    try:
        with open(stats_file, 'r') as stat_file:
            for line in stat_file.readlines():
                try:
                    # search for a match, extract group1 from the match object
                    match = search_pattern_ob.search(line).group(1)
                    match_with_date = "{}_{}".format(match, stat_date)
                except AttributeError:
                    # pdb.set_trace()
                    # Ignoring end of file blank lines
                    pass
                else:
                    try:
                        old_count = enodeb_count_dict[match_with_date]
                    except KeyError:
                        enodeb_count_dict[match_with_date] = 1
                    else:
                        current_count = old_count+1
                        enodeb_count_dict[match_with_date] = current_count
    except (FileNotFoundError, IsADirectoryError):
        print("Stat file {} was not readable".format(stat_file))
    # finally:
    #     as dict is mutable object, so we don't return it.
    #     return enodeb_count_dict


def search_enodeb_n_get_sig_file_count(stats_file, enode_pattern, enodeb_count_dict):
    search_pattern = enode_pattern
    search_pattern_ob = re.compile(search_pattern, re.IGNORECASE)
    try:
        with open(stats_file, 'r') as stat_file:
            for line in stat_file.readlines():
                try:
                    match = search_pattern_ob.search(line).group(1)
                except AttributeError:
                    # pdb.set_trace()
                    # Ignoring end of file blank lines
                    print("No Match found in {}".format(line))
                    pass
                else:
                    try:
                        old_count = enodeb_count_dict[match]
                    except KeyError:
                        enodeb_count_dict[match] = 1
                    else:
                        current_count = old_count+1
                        enodeb_count_dict[match] = current_count
    except (FileNotFoundError, IsADirectoryError):
        print("Stat file {} was not readable")
    # finally:
    #     return enodeb_count_dict


def read_stat_from_a_date_directory(date_directory_path, enodeB_pattern):
    date_directory = os.path.realpath(date_directory_path)
    stat_files_list = os.listdir(date_directory)
    enodeB_count_dict_for_a_date = {}
    for file_name in stat_files_list:
        file_path = os.path.join(date_directory, file_name)
        search_enodeb_n_get_sig_file_count(file_path, enodeB_pattern, enodeB_count_dict_for_a_date)
    return enodeB_count_dict_for_a_date


def write_to_report(enodeb_count_dict_sum: dict, report_file_path):
    w_book = openpyxl.Workbook()
    w_sheet = w_book.active
    w_sheet.cell(1,1,"E_NodeB_Name")
    w_sheet.cell(1,2,"Files_count")
    current_row = 2
    for enodeB, count in enodeb_count_dict_sum.items():
        w_sheet.cell(current_row, 1, enodeB)
        w_sheet.cell(current_row, 2, count)
        current_row +=1
    w_book.save(report_file_path)


def read_from_each_date_directory_n_generate_report(stat_root_dir, enb_pattern, report_directory):
    report_directory = pathlib.Path(report_directory)
    report_directory.mkdir(parents=True, exist_ok=True)
    for date_dir in os.listdir(stat_root_dir):
        report_file_path = os.path.join(report_directory, "{}.xlsx".format(date_dir))
        current_dir_path = os.path.join(stat_root_dir, date_dir)
        file_count_per_enb_sum_dict = read_stat_from_a_date_directory(current_dir_path,enb_pattern)
        write_to_report(file_count_per_enb_sum_dict, report_file_path)


def move_and_read_stat_files(stat_files_directory):
    pass


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("stat_date_dirs_root")
    parser.add_argument("report_dir_path")
    parser.add_argument("-e", "--enbpatt", help="Please provided EnodB Pattern except it is LTE Huawei")
    args = parser.parse_args()
    stat_date_dirs_root = args.stat_date_dirs_root
    report_dir_path = args.report_dir_path
    EnodeB_pattern = args.enbpatt
    if EnodeB_pattern is None:
        EnodeB_pattern = "^([0-9A-Za-z]+)_"
    read_from_each_date_directory_n_generate_report(stat_date_dirs_root, EnodeB_pattern, report_dir_path)
